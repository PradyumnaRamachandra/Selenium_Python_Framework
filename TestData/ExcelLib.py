from Utilities.Config import Config
import openpyxl

def getTestData(sheetname,testcase_name):

    workbook = openpyxl.load_workbook(Config["Test_Data_Path"])
    worksheet = workbook[sheetname]
    data = []
    header = [testcase_name]
    rows = worksheet.max_row
    cols = worksheet.max_column
    for row in range(1, rows + 1):
        if worksheet.cell(row=row, column=1).value == testcase_name:
            for col in range(2, cols):
                header.append(worksheet.cell(row=row - 1, column=col).value)  # to get Header Values
            break

    rows = worksheet.max_row  # Re initialize iterator
    for row in range(1, rows + 1):
        # To pick only if Execute is "Yes"
        if worksheet.cell(row=row, column=1).value == testcase_name and worksheet.cell(row=row,column=cols).value == "Yes":
            temp = {}
            for col in range(2, cols):  # To Eliminate Execute Column
                temp[header[col - 1]] = worksheet.cell(row=row, column=col).value
            data.append(temp)

    return data

def read_locators(sheetname):
    workbook = openpyxl.load_workbook(Config['Locator_File_Path'])
    worksheet = workbook[sheetname]
    locatordict = {}
    rows = worksheet.max_row
    cols = worksheet.max_column
    for row in range(2, rows + 1):
        for col in range(1, cols + 1):
            # print(worksheet.cell(row=row,column=col).value)
            locatordict.update({worksheet.cell(row=row, column=1).value: (worksheet.cell(row=row, column=2).value, worksheet.cell(row=row, column=3).value)})

    return locatordict