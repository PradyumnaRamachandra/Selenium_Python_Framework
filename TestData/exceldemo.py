import openpyxl


#
# workbook=openpyxl.load_workbook("../TestData/TestData.xlsx")
# worksheet=workbook["HomePage"]
# data=[]
# header=["TC"]
# rows=worksheet.max_row
# cols=worksheet.max_column
# for row in range(1,rows+1):
#     if worksheet.cell(row=row,column=1).value=="test_form":
#         for col in range(2,cols):
#             header.append(worksheet.cell(row=row-1,column=col).value)  # to get Header Values
#         break
#
# rows=worksheet.max_row # Re initialize iterator
# for row in range(1,rows+1):
#     #To pick only if Execute is "Yes"
#     if worksheet.cell(row=row,column=1).value=="test_form" and worksheet.cell(row=row,column=cols).value=="Yes":
#         temp = {}
#         for col in range(2,cols): # To Eliminate Execute Column
#             temp[header[col-1]]=worksheet.cell(row=row,column=col).value
#         data.append(temp)
#
#



# #
# loc=[]
# rows=worksheet.max_row
# col=worksheet.max_column
#
# for i in range(1,rows+1):
#     if worksheet.cell(row=i,column=1).value=="Testcase02" and worksheet.cell(row=i,column=col).value=="Yes":
#         temp = {}
#         for j in range(2,col):
#             temp[worksheet.cell(row=1,column=j).value]=worksheet.cell(row=i,column=j).value
#         loc.append(temp)
#
#
# # print(loc)
#
# from TestData.ExcelLib import *
#
# a=getTestData("HomePage","test_formsubmission")
# print(a)

def read_locators():

    workbook = openpyxl.load_workbook("C:\\Users\\pr57\\Desktop\\Selenium_Python\\TestObjects.xlsx")
    worksheet = workbook["HomePage"]
    locatordict = {}
    rows = worksheet.max_row
    cols = worksheet.max_column
    for row in range(2, rows + 1):
        for col in range(1, cols + 1):
            # print(worksheet.cell(row=row,column=col).value)
            locatordict.update({worksheet.cell(row=row, column=1).value: (worksheet.cell(row=row, column=2).value, worksheet.cell(row=row, column=3).value)})

    return locatordict

a=read_locators()
print(a)